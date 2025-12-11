from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models import FinancialStatement
from app.services.financial_statement_service import upsert_financial_rows

LabelMap = Dict[str, str]


def _to_number(value: object) -> Optional[float]:
    try:
        if value is None or value == "":
            return None
        if isinstance(value, str):
            cleaned = value.replace(",", "").strip()
            if cleaned == "":
                return None
            return float(cleaned)
        return float(value)
    except Exception:
        return None


def _find_year_columns(sheet) -> List[int]:
    year_cols: List[int] = []
    for row in sheet.iter_rows(values_only=True):
        candidates: List[int] = []
        for idx, cell in enumerate(row):
            if isinstance(cell, (int, float)) and 2000 <= int(cell) <= 2100:
                candidates.append(idx)
            elif isinstance(cell, str):
                cleaned = cell.replace("年", "").strip()
                if cleaned.isdigit() and 2000 <= int(cleaned) <= 2100:
                    candidates.append(idx)
        if len(candidates) >= 1:
            year_cols = candidates
            break
    if not year_cols:
        # fallback: assume next 3 columns after first data column
        year_cols = list(range(1, 4))
    return year_cols


def _detect_sheet(wb) -> object:
    for name in wb.sheetnames:
        if "入力" in name or "input" in name.lower():
            return wb[name]
    return wb.active


def _find_label_rows(sheet, labels: LabelMap) -> Dict[str, int]:
    positions: Dict[str, int] = {}
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for cell in row:
            if not isinstance(cell, str):
                continue
            cleaned = cell.replace(" ", "").replace("\u3000", "")
            for keyword, field in labels.items():
                if keyword in cleaned and field not in positions:
                    positions[field] = row_idx
    return positions


def _collect_values(sheet, row_idx: int, col_indices: List[int]) -> List[Optional[float]]:
    values: List[Optional[float]] = []
    rows = list(sheet.iter_rows(values_only=True))
    if row_idx >= len(rows):
        return [None] * len(col_indices)
    row = rows[row_idx]
    for col in col_indices:
        if col < len(row):
            values.append(_to_number(row[col]))
        else:
            values.append(None)
    return values


def _build_years(col_indices: List[int]) -> List[int]:
    current_year = datetime.utcnow().year
    return [current_year - idx for idx, _ in enumerate(col_indices)]


def parse_local_benchmark(content: bytes) -> List[Dict[str, Optional[float]]]:
    wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    sheet = _detect_sheet(wb)

    label_map: LabelMap = {
        "売上高": "sales",
        "営業利益": "operating_profit",
        "経常利益": "ordinary_profit",
        "当期純利益": "net_income",
        "減価償却費": "depreciation",
        "従業員": "employees",
        "従業員数": "employees",
        "現金": "cash_and_deposits",
        "現金・預金": "cash_and_deposits",
        "受取手形": "receivables",
        "売掛金": "receivables",
        "棚卸資産": "inventory",
        "負債合計": "total_liabilities",
        "買掛金": "payables",
        "支払手形": "payables",
        "借入金": "borrowings",
        "有利子負債": "borrowings",
        "純資産合計": "equity",
    }

    year_cols = _find_year_columns(sheet)
    label_rows = _find_label_rows(sheet, label_map)
    if not label_rows:
        return []

    year_values = []
    # Try to read explicit year row if available using first label row as anchor
    rows = list(sheet.iter_rows(values_only=True))
    for col in year_cols:
        year = None
        for r in rows[:5]:
            if col < len(r):
                val = r[col]
                if isinstance(val, (int, float)) and 2000 <= int(val) <= 2100:
                    year = int(val)
                    break
                if isinstance(val, str):
                    cleaned = val.replace("年", "").strip()
                    if cleaned.isdigit() and 2000 <= int(cleaned) <= 2100:
                        year = int(cleaned)
                        break
        year_values.append(year)
    if any(year is None for year in year_values):
        year_values = _build_years(year_cols)

    data_by_year: List[Dict[str, Optional[float]]] = [dict(fiscal_year=year_values[i]) for i in range(len(year_cols))]

    for keyword, field in label_map.items():
        if field not in label_rows:
            continue
        row_idx = label_rows[field]
        values = _collect_values(sheet, row_idx, year_cols)
        for idx, val in enumerate(values):
            data_by_year[idx][field] = val

    # Derive current_assets/current_liabilities if possible
    for entry in data_by_year:
        receivables = entry.get("receivables") or 0
        inventory = entry.get("inventory") or 0
        cash = entry.get("cash_and_deposits") or 0
        payables = entry.get("payables") or 0
        total_liabilities = entry.get("total_liabilities")
        if total_liabilities is None and payables:
            entry["total_liabilities"] = payables
        entry.setdefault("current_assets", cash + receivables + inventory)
        entry.setdefault("current_liabilities", payables if payables else None)
    return data_by_year


def upsert_financial_statements(db: Session, company_id: str, content: bytes) -> None:
    rows = parse_local_benchmark(content)
    if not rows:
        return

    upsert_financial_rows(db, company_id, rows)
